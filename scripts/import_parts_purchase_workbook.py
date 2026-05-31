from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from import_workbook_smoke_data import ApiClient, as_text


@dataclass
class PartRow:
    partCode: str
    partBrand: str | None
    partName: str
    specification: str | None
    partCategory: str
    applicableModels: str | None
    source: str
    quantity: int
    unit: str | None
    purchasePrice: str | None
    salePrice: str | None
    settlementPrice: str | None
    remarks: str | None
    inboundDate: str | None


@dataclass
class ImportStats:
    workbook: str = ""
    sourceSheet: str = ""
    sourceRows: int = 0
    groupedParts: int = 0
    partsCreated: int = 0
    partsUpdated: int = 0
    partsReused: int = 0
    skippedRows: list[dict[str, Any]] = field(default_factory=list)


class PartsImportClient(ApiClient):
    def list_parts(self) -> list[dict[str, Any]]:
        return self._request("GET", "/api/parts")

    def create_part(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request("POST", "/api/parts", payload)

    def update_part(self, part_id: int, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request("PUT", f"/api/parts/{part_id}", payload)


def resolve_parts_workbook_path(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
        if not path.exists():
            raise SystemExit(f"workbook not found: {path}")
        return path

    candidates = [
        path for path in Path(r"D:\erp").glob("*.xlsx")
        if not path.name.startswith("~$") and not path.name.startswith("00")
    ]
    if not candidates:
        raise SystemExit("no parts workbook found under D:\\erp")
    return max(candidates, key=lambda path: path.stat().st_size)


def decimal_or_none(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def money(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value.quantize(Decimal("0.01")), "f")


def int_quantity(value: Any) -> int | None:
    number = decimal_or_none(value)
    if number is None:
        return None
    return int(number)


def date_time_iso(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(timespec="seconds")
    if isinstance(value, date):
        return f"{value.isoformat()}T00:00:00"
    text = str(value).strip()
    for separator in (".", "/", "-"):
        parts = text.split(separator)
        if len(parts) == 3 and all(part.isdigit() for part in parts):
            year, month, day = (int(part) for part in parts)
            if 1 <= month <= 12 and 1 <= day <= 31:
                return f"{year:04d}-{month:02d}-{day:02d}T00:00:00"
    return None


def truncate(value: str | None, max_length: int) -> str | None:
    return value[:max_length] if value else None


def classify_part(name: str, specification: str | None, document_type: str | None) -> str:
    joined = " ".join(part for part in (name, specification, document_type) if part)
    rules = [
        ("油", "油品"),
        ("滤", "保养件"),
        ("电池", "电池"),
        ("充电", "充电器"),
        ("轮胎", "轮胎"),
        ("制动", "制动件"),
        ("刹车", "制动件"),
        ("手柄", "操纵件"),
        ("油缸", "液压件"),
        ("液压", "液压件"),
        ("阀", "液压件"),
        ("货叉", "货叉"),
        ("链条", "门架件"),
        ("控制器", "电控件"),
        ("开关", "电控件"),
        ("合格证", "证件"),
    ]
    for needle, category in rules:
        if needle in joined:
            return category
    return "配件"


def build_remark(rows: list[tuple[int, tuple[Any, ...]]]) -> str:
    order_numbers = []
    document_types = []
    notes = []
    for _, row in rows[:5]:
        for index, target in ((2, order_numbers), (3, document_types), (10, notes), (11, notes), (13, notes)):
            text = as_text(row[index]) if len(row) > index else None
            if text and text not in target:
                target.append(text)
    pieces = ["来源：龙工配件采购明细"]
    if order_numbers:
        pieces.append("订单：" + "、".join(order_numbers[:3]))
    if document_types:
        pieces.append("类型：" + "、".join(document_types[:3]))
    if notes:
        pieces.append("备注：" + " / ".join(notes[:3]))
    return truncate("；".join(pieces), 255) or "来源：龙工配件采购明细"


def load_parts(workbook_path: Path, stats: ImportStats) -> list[PartRow]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    stats.workbook = str(workbook_path)
    stats.sourceSheet = sheet.title

    grouped: dict[str, list[tuple[int, tuple[Any, ...]]]] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(as_text(value) for value in row):
            continue
        stats.sourceRows += 1
        code = as_text(row[1]) if len(row) > 1 else None
        name = as_text(row[4]) if len(row) > 4 else None
        quantity = int_quantity(row[7] if len(row) > 7 else None)
        if not code or not name or quantity is None:
            stats.skippedRows.append({"row": row_number, "reason": "missing code, name or quantity"})
            continue
        grouped.setdefault(code, []).append((row_number, row))

    parts: list[PartRow] = []
    for code, rows in grouped.items():
        total_quantity = 0
        latest_date: str | None = None
        latest_row = rows[-1][1]
        weighted_amount = Decimal("0")
        weighted_quantity = Decimal("0")
        for _, row in rows:
            quantity = int_quantity(row[7] if len(row) > 7 else None) or 0
            total_quantity += quantity
            price = decimal_or_none(row[8] if len(row) > 8 else None)
            if price is not None and quantity > 0:
                weighted_amount += price * Decimal(quantity)
                weighted_quantity += Decimal(quantity)
            row_date = date_time_iso(row[0] if len(row) > 0 else None)
            if row_date and (latest_date is None or row_date > latest_date):
                latest_date = row_date
                latest_row = row

        name = as_text(latest_row[4]) or code
        specification = as_text(latest_row[5])
        document_type = as_text(latest_row[3])
        average_price = weighted_amount / weighted_quantity if weighted_quantity else decimal_or_none(latest_row[8])
        parts.append(PartRow(
            partCode=truncate(code, 100) or code,
            partBrand="龙工",
            partName=truncate(name, 100) or code,
            specification=truncate(specification, 100),
            partCategory=truncate(classify_part(name, specification, document_type), 50) or "配件",
            applicableModels=truncate(specification, 255),
            source="采购明细导入",
            quantity=max(0, total_quantity),
            unit=truncate(as_text(latest_row[6]), 20),
            purchasePrice=money(average_price),
            salePrice=None,
            settlementPrice=money(average_price),
            remarks=build_remark(rows),
            inboundDate=latest_date,
        ))

    stats.groupedParts = len(parts)
    workbook.close()
    return sorted(parts, key=lambda part: part.partCode)


def payload_from_part(part: PartRow) -> dict[str, Any]:
    payload = asdict(part)
    return {key: value for key, value in payload.items() if value is not None}


def import_parts(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = resolve_parts_workbook_path(args.workbook)
    stats = ImportStats()
    parts = load_parts(workbook_path, stats)
    client = PartsImportClient(args.base_url, args.username, args.password)
    existing_by_code = {
        part.get("partCode"): part
        for part in client.list_parts()
        if part.get("partCode")
    }

    for part in parts:
        payload = payload_from_part(part)
        existing = existing_by_code.get(part.partCode)
        if existing is None:
            client.create_part(payload)
            stats.partsCreated += 1
            continue
        payload["version"] = existing.get("version")
        comparable_keys = (
            "partBrand", "partName", "specification", "partCategory", "applicableModels",
            "source", "quantity", "unit", "purchasePrice", "settlementPrice", "remarks", "inboundDate",
        )
        changed = any(str(existing.get(key) or "") != str(payload.get(key) or "") for key in comparable_keys)
        if changed:
            client.update_part(existing["id"], payload)
            stats.partsUpdated += 1
        else:
            stats.partsReused += 1

    final_parts = client.list_parts()
    return {
        "baseUrl": args.base_url,
        "stats": asdict(stats),
        "finalCounts": {
            "parts": len(final_parts),
            "partTotalQuantity": sum(part.get("quantity") or 0 for part in final_parts),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Longgong parts purchase workbook through the ERP API.")
    parser.add_argument("--workbook", help="Path to the parts purchase workbook.")
    parser.add_argument("--base-url", default="http://localhost:8080")
    parser.add_argument("--username", default="admin")
    parser.add_argument("--password", default="admin123")
    args = parser.parse_args()
    print(json.dumps(import_parts(args), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
