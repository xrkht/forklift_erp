from __future__ import annotations

import argparse
import json
import os
import re
import ssl
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DATE_TEXT_PATTERN = re.compile(r"(?P<year>\d{4})[-/.年](?P<month>\d{1,2})(?:[-/.月](?P<day>\d{1,2}))?")


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def as_date_iso(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    if len(text) == 6 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}-01"
    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    match = DATE_TEXT_PATTERN.search(text)
    if match:
        year = int(match.group("year"))
        month = int(match.group("month"))
        day = int(match.group("day") or 1)
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def as_datetime_iso(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat(timespec="seconds")
    if isinstance(value, date):
        return f"{value.isoformat()}T00:00:00"
    text = str(value).strip()
    if not text:
        return None
    parsed_date = as_date_iso(text)
    if parsed_date:
        return f"{parsed_date}T00:00:00"
    if len(text) == 10:
        return f"{text}T00:00:00"
    return None


def as_decimal_string(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return format(value.quantize(Decimal("0.01")), "f")
    if isinstance(value, (int, float)):
        return format(Decimal(str(value)).quantize(Decimal("0.01")), "f")
    text = str(value).strip().replace(",", "").replace("元", "")
    if not text or text.startswith("="):
        return None
    try:
        return format(Decimal(text).quantize(Decimal("0.01")), "f")
    except InvalidOperation:
        return None


def parse_bool_from_text(value: Any) -> bool:
    text = as_text(value)
    if not text:
        return False
    normalized = text.replace("（", "(").replace("）", ")")
    false_tokens = ("否", "未", "0", "/", "不")
    return not normalized.startswith(false_tokens)


def parse_inventory_count(value: Any) -> int:
    if isinstance(value, (int, float)):
        return int(value)
    text = as_text(value)
    if not text:
        return 0
    try:
        return int(Decimal(text))
    except InvalidOperation:
        return 0


def join_notes(*values: Any) -> str | None:
    notes = [text for text in (as_text(value) for value in values) if text]
    return " / ".join(notes) or None


def detect_machine_type(name: str | None, specification: str | None, configuration: str | None) -> str:
    joined = " ".join(part for part in [name, specification, configuration] if part)
    if "手动" in joined:
        return "手动叉车"
    if "托盘" in joined and "堆" not in joined:
        return "托盘搬运车"
    if "堆" in joined or "高" in joined:
        return "堆高车"
    if "电" in joined and "内燃" not in joined:
        return "电动叉车"
    if "旧" in joined:
        return "旧车"
    return "内燃叉车"


def resolve_workbook_path(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path)
        if not path.exists():
            raise SystemExit(f"workbook not found: {path}")
        return path

    candidates = sorted(Path(r"D:\erp").glob("00*.xlsx"))
    if not candidates:
        raise SystemExit("no workbook matching D:\\erp\\00*.xlsx")
    return candidates[0]


class ApiClient:
    def __init__(self, base_url: str, username: str, password: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.token = self._login(username, password)

    def _request(self, method: str, path: str, payload: dict[str, Any] | None = None, auth: bool = True) -> Any:
        url = f"{self.base_url}{path}"
        data = None
        headers = {"Content-Type": "application/json"}
        if auth:
            headers["Authorization"] = f"Bearer {self.token}"
        if payload is not None:
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, context=ssl._create_unverified_context()) as response:
                body = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"{method} {path} failed: {exc.code} {body}") from exc
        if body.get("code") != 200:
            raise RuntimeError(f"{method} {path} failed: {body}")
        return body["data"]

    def _login(self, username: str, password: str) -> str:
        data = self._request("POST", "/api/auth/login", {"username": username, "password": password}, auth=False)
        token = data.get("token")
        if not token:
            raise RuntimeError("login succeeded but token missing")
        return token

    def reset_business_data(self) -> dict[str, Any]:
        return self._request("POST", "/api/admin/business-data/reset")

    def create_customer(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request("POST", "/api/customers", payload)

    def create_machine(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request("POST", "/api/inventory", payload)

    def create_vehicle_outbound(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self._request("POST", "/api/outbound-orders/vehicle", payload)


@dataclass
class ImportStats:
    customers: int = 0
    machines: int = 0
    orders: int = 0
    oldMachines: int = 0


def load_rows(workbook_path: Path) -> tuple[list[tuple[Any, ...]], list[tuple[Any, ...]], list[tuple[Any, ...]], dict[str, tuple[Any, ...]]]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    inbound_rows = list(workbook.worksheets[0].iter_rows(min_row=2, values_only=True))
    sales_rows = list(workbook.worksheets[1].iter_rows(min_row=2, values_only=True))
    old_inbound_rows = list(workbook.worksheets[4].iter_rows(min_row=2, values_only=True))
    inbound_by_vehicle = {
        as_text(row[8]): row
        for row in inbound_rows
        if row and len(row) > 8 and as_text(row[8])
    }
    return inbound_rows, sales_rows, old_inbound_rows, inbound_by_vehicle


def build_machine_payload_from_inbound(row: tuple[Any, ...], inventory_count: int) -> dict[str, Any]:
    return {
        "vehicleProductNumber": as_text(row[8]),
        "name": as_text(row[4]),
        "specificationModel": as_text(row[6]),
        "configuration": as_text(row[7]),
        "machineType": detect_machine_type(as_text(row[4]), as_text(row[6]), as_text(row[7])),
        "supplier": as_text(row[2]),
        "warehouseName": as_text(row[24]) or "中山天力",
        "applicationNumber": as_text(row[3]),
        "materialNumber": as_text(row[5]),
        "engineNumber": as_text(row[9]),
        "frameNumber": as_text(row[10]),
        "warrantyCardNumber": as_text(row[11]),
        "manufacturingDate": as_date_iso(row[12]),
        "inboundDate": as_datetime_iso(row[1]),
        "purchasePrice": as_decimal_string(row[13]),
        "settlementPrice": as_decimal_string(row[13]),
        "inventoryCount": inventory_count,
        "destination1": as_text(row[19]),
        "destination2": as_text(row[20]),
        "destination3": as_text(row[21]),
        "destination4": as_text(row[22]),
        "destination5": as_text(row[23]),
        "isSalesReported": as_text(row[15]),
        "salesReportDate": as_date_iso(row[16]),
        "isInvoiceApplied": as_text(row[25]),
        "remarks": join_notes(row[14], row[26]),
    }


def build_machine_payload_from_old_inbound(row: tuple[Any, ...]) -> dict[str, Any]:
    inventory_count = max(1, parse_inventory_count(row[13]))
    return {
        "vehicleProductNumber": as_text(row[6]),
        "name": as_text(row[3]) or "旧车",
        "specificationModel": as_text(row[4]) or "未填写型号",
        "configuration": as_text(row[5]),
        "machineType": detect_machine_type(as_text(row[3]), as_text(row[4]), as_text(row[5])),
        "supplier": "旧车回收",
        "warehouseName": "旧车库存",
        "engineNumber": as_text(row[7]),
        "frameNumber": as_text(row[9]),
        "manufacturingDate": as_date_iso(row[10]),
        "inboundDate": as_datetime_iso(row[1]),
        "salePrice": as_decimal_string(row[8]),
        "inventoryCount": inventory_count,
        "destination1": as_text(row[14]),
        "destination2": as_text(row[15]),
        "destination3": as_text(row[16]),
        "destination4": as_text(row[17]),
        "destination5": as_text(row[18]),
        "remarks": join_notes(row[2], row[11], row[12], row[19]),
    }


def build_customer_payload_from_sales(row: tuple[Any, ...]) -> dict[str, Any]:
    company_name = as_text(row[14])
    return {
        "companyName": company_name,
        "address": as_text(row[15]),
        "contactName": as_text(row[16]) or company_name,
        "contactPhone": as_text(row[17]),
        "taxOrIdNumber": as_text(row[18]),
        "remarks": as_text(row[26]),
    }


def build_vehicle_outbound_payload(machine: dict[str, Any], customer: dict[str, Any], sales_row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "machineId": machine["id"],
        "machineVersion": machine["version"],
        "customerId": customer["id"],
        "salesDate": as_date_iso(sales_row[1]),
        "settlementPrice": as_decimal_string(sales_row[9]),
        "salePrice": as_decimal_string(sales_row[11]),
        "paymentSettled": parse_bool_from_text(sales_row[13]),
        "paymentRemark": as_text(sales_row[12]),
        "salesReported": parse_bool_from_text(sales_row[21]),
        "salesReportDate": as_date_iso(sales_row[22]),
        "invoiceApplied": parse_bool_from_text(sales_row[24]),
        "invoiceApplicationDate": as_date_iso(sales_row[20]),
        "invoiceStatus": as_text(sales_row[19]),
        "invoiceIssuedDate": as_date_iso(sales_row[20]),
        "registrationStatus": as_text(sales_row[25]),
        "contractType": as_text(sales_row[28]),
        "operator": "workbook-smoke",
        "orderRemark": join_notes(sales_row[10], sales_row[26], sales_row[27]),
    }


def select_sold_candidates(
    sales_rows: list[tuple[Any, ...]],
    inbound_by_vehicle: dict[str, tuple[Any, ...]],
    limit: int,
) -> list[tuple[tuple[Any, ...], tuple[Any, ...]]]:
    candidates: list[tuple[tuple[Any, ...], tuple[Any, ...]]] = []
    for sales_row in sales_rows:
        vehicle_number = as_text(sales_row[5]) if len(sales_row) > 5 else None
        customer_name = as_text(sales_row[14]) if len(sales_row) > 14 else None
        inbound_row = inbound_by_vehicle.get(vehicle_number)
        if not vehicle_number or not customer_name or inbound_row is None:
            continue
        candidates.append((sales_row, inbound_row))
        if len(candidates) >= limit:
            break
    return candidates


def main() -> None:
    parser = argparse.ArgumentParser(description="Import a workbook-driven smoke dataset through the ERP API.")
    parser.add_argument("--workbook", help="Path to the workbook. Defaults to the first D:\\erp\\00*.xlsx file.")
    parser.add_argument("--base-url", default="http://localhost:8080")
    parser.add_argument("--username", default=os.environ.get("FORKLIFT_ERP_BOOTSTRAP_ADMIN_USERNAME", "admin"))
    parser.add_argument("--password", default=os.environ.get("FORKLIFT_ERP_BOOTSTRAP_ADMIN_PASSWORD"))
    parser.add_argument("--sold-samples", type=int, default=2)
    parser.add_argument("--stock-samples", type=int, default=3)
    parser.add_argument("--old-stock-samples", type=int, default=1)
    parser.add_argument("--reset-before-import", action="store_true")
    args = parser.parse_args()
    if not args.password:
        parser.error("--password is required, or set FORKLIFT_ERP_BOOTSTRAP_ADMIN_PASSWORD.")

    workbook_path = resolve_workbook_path(args.workbook)
    inbound_rows, sales_rows, old_inbound_rows, inbound_by_vehicle = load_rows(workbook_path)
    client = ApiClient(args.base_url, args.username, args.password)
    stats = ImportStats()
    reset_summary: dict[str, Any] | None = None

    if args.reset_before_import:
        reset_summary = client.reset_business_data()

    created_customers: dict[str, dict[str, Any]] = {}
    created_vehicle_numbers: set[str] = set()

    for sales_row, inbound_row in select_sold_candidates(sales_rows, inbound_by_vehicle, args.sold_samples):
        customer_name = as_text(sales_row[14]) or ""
        if customer_name not in created_customers:
            created_customers[customer_name] = client.create_customer(build_customer_payload_from_sales(sales_row))
            stats.customers += 1
        machine = client.create_machine(build_machine_payload_from_inbound(inbound_row, inventory_count=1))
        created_vehicle_numbers.add(machine["vehicleProductNumber"])
        stats.machines += 1
        client.create_vehicle_outbound(build_vehicle_outbound_payload(machine, created_customers[customer_name], sales_row))
        stats.orders += 1

    for row in inbound_rows:
        vehicle_number = as_text(row[8]) if len(row) > 8 else None
        inventory_count = parse_inventory_count(row[18]) if len(row) > 18 else 0
        if not vehicle_number or vehicle_number in created_vehicle_numbers or inventory_count <= 0:
            continue
        machine = client.create_machine(build_machine_payload_from_inbound(row, inventory_count=inventory_count))
        created_vehicle_numbers.add(machine["vehicleProductNumber"])
        stats.machines += 1
        if stats.machines >= args.sold_samples + args.stock_samples:
            break

    for row in old_inbound_rows:
        vehicle_number = as_text(row[6]) if len(row) > 6 else None
        inventory_count = parse_inventory_count(row[13]) if len(row) > 13 else 0
        if not vehicle_number or vehicle_number in created_vehicle_numbers or inventory_count <= 0:
            continue
        machine = client.create_machine(build_machine_payload_from_old_inbound(row))
        created_vehicle_numbers.add(machine["vehicleProductNumber"])
        stats.machines += 1
        stats.oldMachines += 1
        if stats.oldMachines >= args.old_stock_samples:
            break

    print(json.dumps({
        "baseUrl": args.base_url,
        "workbook": str(workbook_path),
        "resetSummary": reset_summary,
        "customers": stats.customers,
        "machines": stats.machines,
        "orders": stats.orders,
        "oldMachines": stats.oldMachines,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
