from __future__ import annotations

import argparse
import json
import os
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from import_workbook_smoke_data import (
    ApiClient,
    as_date_iso,
    as_datetime_iso,
    as_decimal_string,
    as_text,
    build_customer_payload_from_sales,
    build_machine_payload_from_inbound,
    build_machine_payload_from_old_inbound,
    build_vehicle_outbound_payload,
    detect_machine_type,
    join_notes,
    parse_bool_from_text,
    parse_inventory_count,
    resolve_workbook_path,
)


INVALID_VEHICLE_NUMBERS = {"/", "\\", "-", "--", "0", "none", "null", "n/a", "N/A"}


@dataclass
class FullImportStats:
    workbookRows: dict[str, int] = field(default_factory=dict)
    customersCreated: int = 0
    customersReused: int = 0
    machinesCreated: int = 0
    machinesReused: int = 0
    ordersCreated: int = 0
    ordersSkippedExisting: int = 0
    duplicateVehicleRowsSkipped: int = 0
    generatedVehicleNumbers: int = 0
    skippedRows: list[dict[str, Any]] = field(default_factory=list)


class FullImportClient(ApiClient):
    def list_customers(self) -> list[dict[str, Any]]:
        return self._request("GET", "/api/customers")

    def list_machines(self) -> list[dict[str, Any]]:
        return self._request("GET", "/api/inventory")

    def list_orders(self) -> list[dict[str, Any]]:
        return self._request("GET", "/api/outbound-orders")


def non_empty_rows(sheet: Any) -> list[tuple[int, tuple[Any, ...]]]:
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if any(as_text(value) for value in row):
            rows.append((row_number, row))
    return rows


def clean_vehicle_number(value: Any) -> str | None:
    text = as_text(value)
    if not text:
        return None
    text = text.strip("<> ")
    if text in INVALID_VEHICLE_NUMBERS:
        return None
    return text


def text_at(row: tuple[Any, ...], index: int, max_length: int | None = None) -> str | None:
    value = as_text(row[index]) if len(row) > index else None
    if value and max_length is not None:
        return value[:max_length]
    return value


def decimal_or_zero(*values: Any) -> str:
    for value in values:
        decimal = as_decimal_string(value)
        if decimal is not None:
            return decimal
    return "0.00"


def truncate_payload_text(payload: dict[str, Any], key: str, max_length: int) -> None:
    value = as_text(payload.get(key))
    payload[key] = value[:max_length] if value else None


def normalize_machine_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(payload)
    normalized["vehicleProductNumber"] = as_text(normalized.get("vehicleProductNumber"))
    normalized["name"] = (as_text(normalized.get("name")) or "Workbook vehicle")[:100]
    normalized["specificationModel"] = (as_text(normalized.get("specificationModel")) or "Workbook model")[:100]
    normalized["machineType"] = (
        as_text(normalized.get("machineType"))
        or detect_machine_type(normalized["name"], normalized["specificationModel"], as_text(normalized.get("configuration")))
    )
    truncate_payload_text(normalized, "machineType", 30)
    truncate_payload_text(normalized, "configuration", 500)
    truncate_payload_text(normalized, "supplier", 50)
    truncate_payload_text(normalized, "warehouseName", 100)
    truncate_payload_text(normalized, "applicationNumber", 100)
    truncate_payload_text(normalized, "materialNumber", 100)
    truncate_payload_text(normalized, "engineNumber", 100)
    truncate_payload_text(normalized, "frameNumber", 100)
    truncate_payload_text(normalized, "warrantyCardNumber", 100)
    truncate_payload_text(normalized, "isSalesReported", 10)
    truncate_payload_text(normalized, "isInvoiceApplied", 50)
    truncate_payload_text(normalized, "remarks", 500)
    for key in ("destination1", "destination2", "destination3", "destination4", "destination5"):
        truncate_payload_text(normalized, key, 255)
    return normalized


def generated_vehicle_number(prefix: str, row_number: int, row: tuple[Any, ...]) -> str:
    serial = text_at(row, 0, 20)
    if serial:
        safe_serial = "".join(ch for ch in serial if ch.isalnum())[:20]
        if safe_serial:
            return f"{prefix}-{safe_serial}"
    return f"{prefix}-ROW-{row_number}"


def indexed_by_vehicle(rows: Iterable[tuple[int, tuple[Any, ...]]], vehicle_index: int) -> dict[str, tuple[int, tuple[Any, ...]]]:
    indexed: dict[str, tuple[int, tuple[Any, ...]]] = {}
    for row_number, row in rows:
        vehicle_number = clean_vehicle_number(row[vehicle_index] if len(row) > vehicle_index else None)
        if vehicle_number and vehicle_number not in indexed:
            indexed[vehicle_number] = (row_number, row)
    return indexed


def customer_payload(
    company_name: str,
    address: Any = None,
    contact_name: Any = None,
    contact_phone: Any = None,
    tax_or_id_number: Any = None,
    remarks: Any = None,
) -> dict[str, Any]:
    return {
        "companyName": company_name[:120],
        "address": as_text(address)[:255] if as_text(address) else None,
        "contactName": as_text(contact_name)[:80] if as_text(contact_name) else company_name[:80],
        "contactPhone": as_text(contact_phone)[:50] if as_text(contact_phone) else None,
        "taxOrIdNumber": as_text(tax_or_id_number)[:100] if as_text(tax_or_id_number) else None,
        "remarks": as_text(remarks)[:500] if as_text(remarks) else None,
    }


def fallback_customer_name(prefix: str, vehicle_number: str, *values: Any) -> str:
    for value in values:
        text = as_text(value)
        if text:
            return text[:120]
    return f"{prefix}-{vehicle_number}"[:120]


def machine_from_sales_row(row: tuple[Any, ...], vehicle_number: str, source: str) -> dict[str, Any]:
    name = text_at(row, 2, 100) or "Workbook vehicle"
    specification = text_at(row, 3, 100) or text_at(row, 4, 100) or "Workbook model"
    configuration = text_at(row, 4, 500)
    return {
        "vehicleProductNumber": vehicle_number,
        "name": name,
        "specificationModel": specification,
        "configuration": configuration,
        "machineType": detect_machine_type(name, specification, configuration),
        "supplier": text_at(row, 27, 50) or source,
        "warehouseName": source,
        "engineNumber": text_at(row, 6, 100),
        "frameNumber": text_at(row, 7, 100),
        "warrantyCardNumber": text_at(row, 8, 100),
        "inboundDate": as_datetime_iso(row[1] if len(row) > 1 else None),
        "purchasePrice": as_decimal_string(row[9] if len(row) > 9 else None),
        "settlementPrice": as_decimal_string(row[9] if len(row) > 9 else None),
        "salePrice": as_decimal_string(row[11] if len(row) > 11 else None),
        "inventoryCount": 1,
        "stockStatus": "IN_STOCK",
        "isSalesReported": text_at(row, 21, 10),
        "salesReportDate": as_date_iso(row[22] if len(row) > 22 else None),
        "isInvoiceApplied": text_at(row, 24, 50),
        "remarks": join_notes(source, row[10] if len(row) > 10 else None, row[26] if len(row) > 26 else None),
    }


def machine_from_other_brand_sales(row: tuple[Any, ...], vehicle_number: str) -> dict[str, Any]:
    name = text_at(row, 2, 100) or "Other brand vehicle"
    specification = text_at(row, 4, 100) or "Other brand model"
    configuration = text_at(row, 5, 500)
    brand = text_at(row, 3, 50)
    return {
        "vehicleProductNumber": vehicle_number,
        "name": name,
        "specificationModel": specification,
        "configuration": configuration,
        "machineType": detect_machine_type(name, specification, configuration),
        "supplier": brand or "Other brand",
        "warehouseName": "Other brand sales",
        "engineNumber": text_at(row, 7, 100),
        "frameNumber": text_at(row, 8, 100),
        "warrantyCardNumber": text_at(row, 9, 100),
        "inboundDate": as_datetime_iso(row[1] if len(row) > 1 else None),
        "settlementPrice": as_decimal_string(row[10] if len(row) > 10 else None),
        "salePrice": as_decimal_string(row[10] if len(row) > 10 else None),
        "inventoryCount": 1,
        "stockStatus": "IN_STOCK",
        "remarks": join_notes("Other brand sales", brand, row[20] if len(row) > 20 else None),
    }


def machine_from_old_sales(row: tuple[Any, ...], vehicle_number: str) -> dict[str, Any]:
    name = text_at(row, 2, 100) or "Used vehicle"
    specification = text_at(row, 3, 100) or "Used vehicle model"
    configuration = text_at(row, 4, 500)
    brand = text_at(row, 21, 50)
    return {
        "vehicleProductNumber": vehicle_number,
        "name": name,
        "specificationModel": specification,
        "configuration": configuration,
        "machineType": detect_machine_type(name, specification, configuration),
        "supplier": brand or "Used vehicle",
        "warehouseName": "Used vehicle sales",
        "engineNumber": text_at(row, 6, 100),
        "frameNumber": text_at(row, 7, 100),
        "warrantyCardNumber": text_at(row, 8, 100),
        "inboundDate": as_datetime_iso(row[1] if len(row) > 1 else None),
        "settlementPrice": as_decimal_string(row[11] if len(row) > 11 else None),
        "salePrice": as_decimal_string(row[11] if len(row) > 11 else None),
        "inventoryCount": 1,
        "stockStatus": "IN_STOCK",
        "remarks": join_notes("Used vehicle sales", brand, row[20] if len(row) > 20 else None),
    }


def outbound_payload_from_other_brand(machine: dict[str, Any], customer: dict[str, Any], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "machineId": machine["id"],
        "machineVersion": machine["version"],
        "customerId": customer["id"],
        "salesDate": as_date_iso(row[1] if len(row) > 1 else None),
        "settlementPrice": decimal_or_zero(row[10] if len(row) > 10 else None),
        "salePrice": as_decimal_string(row[10] if len(row) > 10 else None),
        "paymentSettled": parse_bool_from_text(row[12] if len(row) > 12 else None),
        "paymentRemark": text_at(row, 11, 500),
        "invoiceStatus": text_at(row, 18, 100),
        "invoiceIssuedDate": as_date_iso(row[19] if len(row) > 19 else None),
        "operator": "workbook-full-import",
        "orderRemark": join_notes("Other brand sales", row[20] if len(row) > 20 else None),
    }


def outbound_payload_from_old_sales(machine: dict[str, Any], customer: dict[str, Any], row: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "machineId": machine["id"],
        "machineVersion": machine["version"],
        "customerId": customer["id"],
        "salesDate": as_date_iso(row[1] if len(row) > 1 else None),
        "settlementPrice": decimal_or_zero(row[11] if len(row) > 11 else None),
        "salePrice": as_decimal_string(row[11] if len(row) > 11 else None),
        "invoiceStatus": text_at(row, 18, 100),
        "invoiceIssuedDate": as_date_iso(row[19] if len(row) > 19 else None),
        "operator": "workbook-full-import",
        "orderRemark": join_notes(
            "Used vehicle sales",
            f"quantity={text_at(row, 10, 20)}" if text_at(row, 10, 20) else None,
            row[20] if len(row) > 20 else None,
        ),
    }


def refresh_indexes(client: FullImportClient) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], set[str]]:
    customers = {
        customer["companyName"]: customer
        for customer in client.list_customers()
        if customer.get("companyName")
    }
    machines = {
        machine["vehicleProductNumber"]: machine
        for machine in client.list_machines()
        if machine.get("vehicleProductNumber")
    }
    order_vehicle_numbers = {
        order["resourceCode"]
        for order in client.list_orders()
        if order.get("resourceType") == "MACHINE" and order.get("resourceCode")
    }
    return customers, machines, order_vehicle_numbers


def create_or_get_customer(
    client: FullImportClient,
    customers: dict[str, dict[str, Any]],
    payload: dict[str, Any],
    stats: FullImportStats,
) -> dict[str, Any]:
    company_name = payload["companyName"]
    existing = customers.get(company_name)
    if existing:
        stats.customersReused += 1
        return existing
    created = client.create_customer(payload)
    customers[company_name] = created
    stats.customersCreated += 1
    return created


def create_or_get_machine(
    client: FullImportClient,
    machines: dict[str, dict[str, Any]],
    payload: dict[str, Any],
    stats: FullImportStats,
) -> dict[str, Any]:
    payload = normalize_machine_payload(payload)
    vehicle_number = payload["vehicleProductNumber"]
    existing = machines.get(vehicle_number)
    if existing:
        stats.machinesReused += 1
        return existing
    created = client.create_machine(payload)
    machines[vehicle_number] = created
    stats.machinesCreated += 1
    return created


def create_order_if_missing(
    client: FullImportClient,
    machines: dict[str, dict[str, Any]],
    order_vehicle_numbers: set[str],
    vehicle_number: str,
    payload: dict[str, Any],
    stats: FullImportStats,
) -> None:
    if vehicle_number in order_vehicle_numbers:
        stats.ordersSkippedExisting += 1
        return
    machine = machines[vehicle_number]
    if (machine.get("inventoryCount") or 0) < 1:
        stats.skippedRows.append({
            "reason": "machine has no stock for outbound",
            "vehicleProductNumber": vehicle_number,
        })
        return
    client.create_vehicle_outbound(payload)
    order_vehicle_numbers.add(vehicle_number)
    machine["inventoryCount"] = max(0, (machine.get("inventoryCount") or 0) - 1)
    machine["stockStatus"] = "OUTBOUND" if machine["inventoryCount"] == 0 else "IN_STOCK"
    stats.ordersCreated += 1


def import_workbook(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = resolve_workbook_path(args.workbook)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)

    inbound_rows = non_empty_rows(workbook.worksheets[0])
    sales_rows = non_empty_rows(workbook.worksheets[1])
    auto_report_rows = non_empty_rows(workbook.worksheets[2])
    other_brand_rows = non_empty_rows(workbook.worksheets[3])
    old_inbound_rows = non_empty_rows(workbook.worksheets[4])
    old_sales_rows = non_empty_rows(workbook.worksheets[5])

    stats = FullImportStats(workbookRows={
        "inbound": len(inbound_rows),
        "sales": len(sales_rows),
        "autoReport": len(auto_report_rows),
        "otherBrandSales": len(other_brand_rows),
        "oldInbound": len(old_inbound_rows),
        "oldSales": len(old_sales_rows),
    })

    inbound_by_vehicle = indexed_by_vehicle(inbound_rows, 8)
    old_inbound_by_vehicle = indexed_by_vehicle(old_inbound_rows, 6)
    stats.duplicateVehicleRowsSkipped += len([
        1 for _, row in inbound_rows
        if clean_vehicle_number(row[8] if len(row) > 8 else None)
    ]) - len(inbound_by_vehicle)

    client = FullImportClient(args.base_url, args.username, args.password)
    if args.reset_before_import:
        reset_summary = client.reset_business_data()
    else:
        reset_summary = None
    customers, machines, order_vehicle_numbers = refresh_indexes(client)

    seen_sale_vehicles: set[str] = set()
    for row_number, sales_row in sales_rows:
        vehicle_number = clean_vehicle_number(sales_row[5] if len(sales_row) > 5 else None)
        if not vehicle_number:
            continue
        if vehicle_number in seen_sale_vehicles:
            stats.duplicateVehicleRowsSkipped += 1
            continue
        seen_sale_vehicles.add(vehicle_number)
        inbound_match = inbound_by_vehicle.get(vehicle_number)
        if inbound_match:
            _, inbound_row = inbound_match
            machine_payload = build_machine_payload_from_inbound(inbound_row, inventory_count=1)
        else:
            machine_payload = machine_from_sales_row(sales_row, vehicle_number, "Workbook sales")
        machine_payload["vehicleProductNumber"] = vehicle_number
        machine_payload["stockStatus"] = "IN_STOCK"
        machine = create_or_get_machine(client, machines, machine_payload, stats)

        customer_name = fallback_customer_name(
            "Workbook-customer",
            vehicle_number,
            sales_row[14] if len(sales_row) > 14 else None,
            inbound_match[1][24] if inbound_match and len(inbound_match[1]) > 24 else None,
            inbound_match[1][19] if inbound_match and len(inbound_match[1]) > 19 else None,
        )
        customer = create_or_get_customer(
            client,
            customers,
            build_customer_payload_from_sales(sales_row)
            if text_at(sales_row, 14)
            else customer_payload(customer_name, remarks="Generated from workbook row without customer name"),
            stats,
        )
        outbound_payload = build_vehicle_outbound_payload(machine, customer, sales_row)
        outbound_payload["settlementPrice"] = decimal_or_zero(sales_row[9] if len(sales_row) > 9 else None)
        create_order_if_missing(client, machines, order_vehicle_numbers, vehicle_number, outbound_payload, stats)

    for row_number, inbound_row in inbound_rows:
        vehicle_number = clean_vehicle_number(inbound_row[8] if len(inbound_row) > 8 else None)
        if not vehicle_number or vehicle_number in machines:
            continue
        inventory_count = max(0, parse_inventory_count(inbound_row[18] if len(inbound_row) > 18 else None))
        payload = build_machine_payload_from_inbound(inbound_row, inventory_count=inventory_count)
        payload["vehicleProductNumber"] = vehicle_number
        payload["stockStatus"] = "IN_STOCK" if inventory_count > 0 else "OUTBOUND"
        create_or_get_machine(client, machines, payload, stats)

    seen_other_vehicles: set[str] = set()
    for row_number, row in other_brand_rows:
        vehicle_number = clean_vehicle_number(row[6] if len(row) > 6 else None)
        if not vehicle_number and text_at(row, 13):
            vehicle_number = generated_vehicle_number("OTHER-SALE", row_number, row)
            stats.generatedVehicleNumbers += 1
        if not vehicle_number:
            stats.skippedRows.append({"sheet": "otherBrandSales", "row": row_number, "reason": "missing vehicle number"})
            continue
        if vehicle_number in seen_other_vehicles:
            stats.duplicateVehicleRowsSkipped += 1
            continue
        seen_other_vehicles.add(vehicle_number)
        machine = create_or_get_machine(client, machines, machine_from_other_brand_sales(row, vehicle_number), stats)
        customer_name = fallback_customer_name("Other-brand-customer", vehicle_number, row[13] if len(row) > 13 else None)
        customer = create_or_get_customer(
            client,
            customers,
            customer_payload(
                customer_name,
                row[14] if len(row) > 14 else None,
                row[15] if len(row) > 15 else None,
                row[16] if len(row) > 16 else None,
                row[17] if len(row) > 17 else None,
                row[20] if len(row) > 20 else None,
            ),
            stats,
        )
        create_order_if_missing(
            client,
            machines,
            order_vehicle_numbers,
            vehicle_number,
            outbound_payload_from_other_brand(machine, customer, row),
            stats,
        )

    seen_old_sale_vehicles: set[str] = set()
    for row_number, row in old_sales_rows:
        vehicle_number = clean_vehicle_number(row[5] if len(row) > 5 else None)
        if not vehicle_number and text_at(row, 13):
            vehicle_number = generated_vehicle_number("OLD-SALE", row_number, row)
            stats.generatedVehicleNumbers += 1
        if not vehicle_number:
            stats.skippedRows.append({"sheet": "oldSales", "row": row_number, "reason": "missing vehicle number"})
            continue
        if vehicle_number in seen_old_sale_vehicles:
            stats.duplicateVehicleRowsSkipped += 1
            continue
        seen_old_sale_vehicles.add(vehicle_number)
        old_inbound_match = old_inbound_by_vehicle.get(vehicle_number)
        if old_inbound_match:
            payload = build_machine_payload_from_old_inbound(old_inbound_match[1])
            payload["vehicleProductNumber"] = vehicle_number
            payload["inventoryCount"] = 1
            payload["stockStatus"] = "IN_STOCK"
        else:
            payload = machine_from_old_sales(row, vehicle_number)
        machine = create_or_get_machine(client, machines, payload, stats)
        customer_name = fallback_customer_name("Used-vehicle-customer", vehicle_number, row[13] if len(row) > 13 else None)
        customer = create_or_get_customer(
            client,
            customers,
            customer_payload(
                customer_name,
                row[14] if len(row) > 14 else None,
                row[15] if len(row) > 15 else None,
                row[16] if len(row) > 16 else None,
                row[17] if len(row) > 17 else None,
                row[20] if len(row) > 20 else None,
            ),
            stats,
        )
        create_order_if_missing(
            client,
            machines,
            order_vehicle_numbers,
            vehicle_number,
            outbound_payload_from_old_sales(machine, customer, row),
            stats,
        )

    for row_number, row in old_inbound_rows:
        vehicle_number = clean_vehicle_number(row[6] if len(row) > 6 else None)
        if not vehicle_number or vehicle_number in machines:
            continue
        inventory_count = max(0, parse_inventory_count(row[13] if len(row) > 13 else None))
        payload = build_machine_payload_from_old_inbound(row)
        payload["vehicleProductNumber"] = vehicle_number
        payload["inventoryCount"] = inventory_count
        payload["stockStatus"] = "IN_STOCK" if inventory_count > 0 else "OUTBOUND"
        create_or_get_machine(client, machines, payload, stats)

    final_customers, final_machines, final_order_vehicle_numbers = refresh_indexes(client)
    final_machine_values = list(final_machines.values())
    return {
        "baseUrl": args.base_url,
        "workbook": str(workbook_path),
        "resetSummary": reset_summary,
        "stats": asdict(stats),
        "finalCounts": {
            "customers": len(final_customers),
            "vehicles": len(final_machines),
            "orders": len(final_order_vehicle_numbers),
            "inStockVehicles": sum(1 for machine in final_machine_values if (machine.get("inventoryCount") or 0) > 0),
            "soldOrZeroStockVehicles": sum(1 for machine in final_machine_values if (machine.get("inventoryCount") or 0) == 0),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the vehicle workbook through the ERP API.")
    parser.add_argument("--workbook", help="Path to the workbook. Defaults to the first D:\\erp\\00*.xlsx file.")
    parser.add_argument("--base-url", default="http://localhost:8080")
    parser.add_argument("--username", default=os.environ.get("FORKLIFT_ERP_BOOTSTRAP_ADMIN_USERNAME", "admin"))
    parser.add_argument("--password", default=os.environ.get("FORKLIFT_ERP_BOOTSTRAP_ADMIN_PASSWORD"))
    parser.add_argument("--reset-before-import", action="store_true")
    args = parser.parse_args()
    if not args.password:
        parser.error("--password is required, or set FORKLIFT_ERP_BOOTSTRAP_ADMIN_PASSWORD.")

    print(json.dumps(import_workbook(args), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
